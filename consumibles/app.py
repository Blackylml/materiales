from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from werkzeug.utils import secure_filename
import os
import datetime
import uuid
import pyodbc
import paramiko
import firebirdsql
import sshtunnel
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from io import BytesIO
from config import Config

app = Flask(__name__)
app.config.from_object(Config)

# Asegurar que existe el directorio de uploads
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Configurar la base de datos
db = SQLAlchemy(app)

# Definir modelos
class ReciboMaterial(db.Model):
    __tablename__ = 'recibos_consumibles'
    
    id = db.Column(db.Integer, primary_key=True)
    idcode = db.Column(db.String(50))
    fecha = db.Column(db.Date, default=datetime.datetime.now)
    orden_compra = db.Column(db.String(100))
    proveedor = db.Column(db.String(200))
    num_remision = db.Column(db.String(100))
    cantidad = db.Column(db.Float)
    tipo = db.Column(db.String(100))
    descripcion_material = db.Column(db.String(500))
    grado_acero = db.Column(db.String(100))
    num_placa = db.Column(db.String(100))
    num_colada = db.Column(db.String(100))
    num_certificado = db.Column(db.String(100))
    ot = db.Column(db.String(100))
    cliente = db.Column(db.String(200))
    estatus = db.Column(db.String(50))
    reporte_focc03 = db.Column(db.String(100))
    idordencompra = db.Column(db.Integer)
    procedencia = db.Column(db.String(100))
    archivo = db.Column(db.String(255))
    fecha_creacion = db.Column(db.DateTime, default=datetime.datetime.now)
    fecha_modificacion = db.Column(db.DateTime, default=datetime.datetime.now, onupdate=datetime.datetime.now)

# Funciones de utilidad para bases de datos
def get_sqlserver_conn():
    """Conexión a SQL Server local"""
    conn_str = f"DRIVER={Config.SQLSERVER_LOCAL['driver']};SERVER={Config.SQLSERVER_LOCAL['server']};DATABASE={Config.SQLSERVER_LOCAL['database']};UID={Config.SQLSERVER_LOCAL['username']};PWD={Config.SQLSERVER_LOCAL['password']}"
    return pyodbc.connect(conn_str)

def get_sqlserver_prod_conn():
    """Conexión a SQL Server de producción"""
    conn_str = f"DRIVER={Config.SQLSERVER_PROD['driver']};SERVER={Config.SQLSERVER_PROD['server']};DATABASE={Config.SQLSERVER_PROD['database']};UID={Config.SQLSERVER_PROD['username']};PWD={Config.SQLSERVER_PROD['password']}"
    return pyodbc.connect(conn_str)

def get_firebird_conn():
    """Establece túnel SSH y conexión a Firebird"""
    try:
        # Verificar si ya existe un túnel
        ssh_config = Config.SSH_CONFIG
        tunnel = sshtunnel.SSHTunnelForwarder(
            (ssh_config['host'], ssh_config['port']),
            ssh_username=ssh_config['username'],
            ssh_password=ssh_config['password'],
            remote_bind_address=(ssh_config['remote_host'], ssh_config['remote_port']),
            local_bind_address=('127.0.0.1', ssh_config['local_port'])
        )
        
        tunnel.start()
        
        firebird_config = Config.FIREBIRD_CONFIG
        conn = firebirdsql.connect(
            host='localhost',
            database=firebird_config['database'],
            user=firebird_config['user'],
            password=firebird_config['password'],
            charset=firebird_config['charset'],
            port=tunnel.local_bind_port
        )
        
        return {'status': True, 'conn': conn, 'tunnel': tunnel, 'message': 'Conexión exitosa'}
    except Exception as e:
        if 'tunnel' in locals() and tunnel.is_active:
            tunnel.close()
        return {'status': False, 'message': f'Error en conexión: {str(e)}'}

def get_procedencias():
    """Obtiene lista de procedencias de SQL Server"""
    try:
        conn = get_sqlserver_prod_conn()
        cursor = conn.cursor()
        cursor.execute("SELECT idprocedencia as id, descripcion FROM tbprocedenciacalidad ORDER BY descripcion")
        procedencias = [{'id': row[0], 'descripcion': row[1]} for row in cursor.fetchall()]
        conn.close()
        return procedencias
    except Exception as e:
        print(f"Error al obtener procedencias: {str(e)}")
        return []

# Rutas de la aplicación
@app.route('/')
def index():
    filtro = request.args.get('filtro', '')
    
    # Obtener recibos con filtro opcional
    if filtro:
        query = ReciboMaterial.query.filter(
            db.or_(
                ReciboMaterial.idcode.ilike(f'%{filtro}%'),
                ReciboMaterial.orden_compra.ilike(f'%{filtro}%'),
                ReciboMaterial.proveedor.ilike(f'%{filtro}%'),
                ReciboMaterial.descripcion_material.ilike(f'%{filtro}%'),
                ReciboMaterial.cliente.ilike(f'%{filtro}%')
            )
        )
    else:
        query = ReciboMaterial.query
    
    recibos = query.order_by(ReciboMaterial.fecha_creacion.desc()).all()
    procedencias = get_procedencias()
    
    today = datetime.datetime.now().strftime('%Y-%m-%d')
    
    return render_template('index.html', recibos=recibos, filtro=filtro, procedencias=procedencias, today=today)

@app.route('/guardar_recibo', methods=['POST'])
def guardar_recibo():
    try:
        accion = request.form.get('accion')
        
        if accion not in ['nuevo', 'editar']:
            flash('Acción no válida', 'danger')
            return redirect(url_for('index'))
        
        # Obtener datos del formulario
        datos = {
            'idcode': request.form.get('idcode', ''),
            'fecha': request.form.get('fecha', datetime.datetime.now().strftime('%Y-%m-%d')),
            'orden_compra': request.form.get('orden_compra', ''),
            'proveedor': request.form.get('proveedor', ''),
            'num_remision': request.form.get('num_remision', ''),
            'cantidad': request.form.get('cantidad', ''),
            'tipo': request.form.get('tipo', ''),
            'descripcion_material': request.form.get('descripcion_material', ''),
            'grado_acero': request.form.get('grado_acero', ''),
            'num_placa': request.form.get('num_placa', ''),
            'num_colada': request.form.get('num_colada', ''),
            'num_certificado': request.form.get('num_certificado', ''),
            'ot': request.form.get('ot', ''),
            'cliente': request.form.get('cliente', ''),
            'estatus': request.form.get('estatus', ''),
            'reporte_focc03': request.form.get('reporte_focc03', ''),
            'procedencia': request.form.get('procedencia', '')
        }
        
        # Procesar archivo adjunto
        archivo = request.files.get('archivo')
        nombre_archivo = None
        
        if archivo and archivo.filename:
            # Generar nombre único para el archivo
            extension = os.path.splitext(archivo.filename)[1]
            nombre_archivo = f"{uuid.uuid4().hex}{extension}"
            ruta_completa = os.path.join(app.config['UPLOAD_FOLDER'], nombre_archivo)
            
            # Guardar archivo
            archivo.save(ruta_completa)
        
        # Crear o actualizar recibo
        if accion == 'nuevo':
            recibo = ReciboMaterial(
                idcode=datos['idcode'],
                fecha=datetime.datetime.strptime(datos['fecha'], '%Y-%m-%d').date() if datos['fecha'] else None,
                orden_compra=datos['orden_compra'],
                proveedor=datos['proveedor'],
                num_remision=datos['num_remision'],
                cantidad=float(datos['cantidad']) if datos['cantidad'] else None,
                tipo=datos['tipo'],
                descripcion_material=datos['descripcion_material'],
                grado_acero=datos['grado_acero'],
                num_placa=datos['num_placa'],
                num_colada=datos['num_colada'],
                num_certificado=datos['num_certificado'],
                ot=datos['ot'],
                cliente=datos['cliente'],
                estatus=datos['estatus'],
                reporte_focc03=datos['reporte_focc03'],
                procedencia=datos['procedencia'],
                archivo=nombre_archivo
            )
            db.session.add(recibo)
            mensaje = 'Recibo creado correctamente'
        else:
            # Editar recibo existente
            id_recibo = request.form.get('id')
            recibo = ReciboMaterial.query.get_or_404(id_recibo)
            
            # Actualizar campos
            recibo.idcode = datos['idcode']
            recibo.fecha = datetime.datetime.strptime(datos['fecha'], '%Y-%m-%d').date() if datos['fecha'] else None
            recibo.orden_compra = datos['orden_compra']
            recibo.proveedor = datos['proveedor']
            recibo.num_remision = datos['num_remision']
            recibo.cantidad = float(datos['cantidad']) if datos['cantidad'] else None
            recibo.tipo = datos['tipo']
            recibo.descripcion_material = datos['descripcion_material']
            recibo.grado_acero = datos['grado_acero']
            recibo.num_placa = datos['num_placa']
            recibo.num_colada = datos['num_colada']
            recibo.num_certificado = datos['num_certificado']
            recibo.ot = datos['ot']
            recibo.cliente = datos['cliente']
            recibo.estatus = datos['estatus']
            recibo.reporte_focc03 = datos['reporte_focc03']
            recibo.procedencia = datos['procedencia']
            
            # Actualizar archivo solo si hay uno nuevo
            if nombre_archivo:
                # Eliminar archivo anterior si existe
                if recibo.archivo:
                    ruta_anterior = os.path.join(app.config['UPLOAD_FOLDER'], recibo.archivo)
                    if os.path.exists(ruta_anterior):
                        os.remove(ruta_anterior)
                
                recibo.archivo = nombre_archivo
            
            mensaje = 'Recibo actualizado correctamente'
        
        db.session.commit()
        flash(mensaje, 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'danger')
    
    return redirect(url_for('index'))

@app.route('/obtener_recibo/<int:id>')
def obtener_recibo(id):
    recibo = ReciboMaterial.query.get_or_404(id)
    
    # Convertir a diccionario
    data = {
        'id': recibo.id,
        'idcode': recibo.idcode,
        'fecha': recibo.fecha.strftime('%Y-%m-%d') if recibo.fecha else '',
        'orden_compra': recibo.orden_compra,
        'proveedor': recibo.proveedor,
        'num_remision': recibo.num_remision,
        'cantidad': recibo.cantidad,
        'tipo': recibo.tipo,
        'descripcion_material': recibo.descripcion_material,
        'grado_acero': recibo.grado_acero,
        'num_placa': recibo.num_placa,
        'num_colada': recibo.num_colada,
        'num_certificado': recibo.num_certificado,
        'ot': recibo.ot,
        'cliente': recibo.cliente,
        'estatus': recibo.estatus,
        'reporte_focc03': recibo.reporte_focc03,
        'procedencia': recibo.procedencia,
        'archivo': recibo.archivo
    }
    
    return jsonify({'status': 'success', 'recibo': data})

@app.route('/detalles_recibo/<int:id>')
def detalles_recibo(id):
    recibo = ReciboMaterial.query.get_or_404(id)
    
    # Obtener nombre de procedencia si existe
    nombre_procedencia = ""
    if recibo.procedencia:
        try:
            conn = get_sqlserver_prod_conn()
            cursor = conn.cursor()
            cursor.execute("SELECT descripcion FROM tbprocedenciacalidad WHERE idprocedencia = ?", recibo.procedencia)
            row = cursor.fetchone()
            if row:
                nombre_procedencia = row[0]
            conn.close()
        except Exception as e:
            print(f"Error al obtener procedencia: {str(e)}")
    
    return render_template('detalles_recibo.html', recibo=recibo, nombre_procedencia=nombre_procedencia)

@app.route('/descargar_archivo/<int:id>')
def descargar_archivo(id):
    recibo = ReciboMaterial.query.get_or_404(id)
    
    if not recibo.archivo:
        flash('Archivo no encontrado', 'danger')
        return redirect(url_for('index'))
    
    ruta_archivo = os.path.join(app.config['UPLOAD_FOLDER'], recibo.archivo)
    
    if not os.path.exists(ruta_archivo):
        flash('El archivo físico no existe', 'danger')
        return redirect(url_for('index'))
    
    return send_file(ruta_archivo, as_attachment=True)

@app.route('/exportar_excel')
def exportar_excel():
    try:
        # Determinar si exportar todos o seleccionados
        todos = request.args.get('todos') == '1'
        ids = request.args.get('ids')
        
        if not todos and not ids:
            flash('No se especificaron IDs para exportar', 'danger')
            return redirect(url_for('index'))
        
        # Obtener recibos
        if todos:
            recibos = ReciboMaterial.query.order_by(ReciboMaterial.fecha_creacion.desc()).all()
        else:
            ids_list = [int(id) for id in ids.split(',')]
            recibos = ReciboMaterial.query.filter(ReciboMaterial.id.in_(ids_list)).order_by(ReciboMaterial.fecha_creacion.desc()).all()
        
        if not recibos:
            flash('No hay recibos para exportar', 'danger')
            return redirect(url_for('index'))
        
        # Crear libro de Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Recibos de Material"
        
        # Estilo para encabezados
        header_font = Font(color="FFFFFF", bold=True)
        header_fill = PatternFill(start_color="DC0000", end_color="DC0000", fill_type="solid")
        
        # Definir encabezados
        headers = [
            'ID Code', 'Fecha', 'Orden de Compra', 'Proveedor', 'Número de Remisión',
            'Cantidad', 'Tipo', 'Descripción del Material', 'Grado de Acero', 'Número de Placa',
            'Número de Colada', 'Número de Certificado', 'OT', 'Cliente', 'Estatus',
            'Reporte FO-CC-03', 'Procedencia'
        ]
        
        # Escribir encabezados
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Obtener conexión para procedencias
        conn = None
        try:
            conn = get_sqlserver_prod_conn()
            cursor = conn.cursor()
            
            # Escribir datos
            for row_num, recibo in enumerate(recibos, 2):
                ws.cell(row=row_num, column=1, value=recibo.idcode)
                ws.cell(row=row_num, column=2, value=recibo.fecha.strftime('%d/%m/%Y') if recibo.fecha else '')
                ws.cell(row=row_num, column=3, value=recibo.orden_compra)
                ws.cell(row=row_num, column=4, value=recibo.proveedor)
                ws.cell(row=row_num, column=5, value=recibo.num_remision)
                ws.cell(row=row_num, column=6, value=recibo.cantidad)
                ws.cell(row=row_num, column=7, value=recibo.tipo)
                ws.cell(row=row_num, column=8, value=recibo.descripcion_material)
                ws.cell(row=row_num, column=9, value=recibo.grado_acero)
                ws.cell(row=row_num, column=10, value=recibo.num_placa)
                ws.cell(row=row_num, column=11, value=recibo.num_colada)
                ws.cell(row=row_num, column=12, value=recibo.num_certificado)
                ws.cell(row=row_num, column=13, value=recibo.ot)
                ws.cell(row=row_num, column=14, value=recibo.cliente)
                ws.cell(row=row_num, column=15, value=recibo.estatus)
                ws.cell(row=row_num, column=16, value=recibo.reporte_focc03)
                
                # Obtener nombre de procedencia
                nombre_procedencia = ''
                if recibo.procedencia:
                    cursor.execute("SELECT descripcion FROM tbprocedenciacalidad WHERE idprocedencia = ?", recibo.procedencia)
                    row = cursor.fetchone()
                    if row:
                        nombre_procedencia = row[0]
                
                ws.cell(row=row_num, column=17, value=nombre_procedencia)
        finally:
            if conn:
                conn.close()
        
        # Ajustar anchos de columna
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        # Crear buffer para el archivo
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Nombre del archivo
        filename = f"Recibos_Material_{datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'Error al exportar: {str(e)}', 'danger')
        return redirect(url_for('index'))

@app.route('/exportar_reporte_focc03')
def exportar_reporte_focc03():
    try:
        # Obtener el ID del reporte
        reporte = request.args.get('reporte', '')
        
        if not reporte:
            flash('Reporte no especificado', 'danger')
            return redirect(url_for('index'))
        
        # Obtener recibos con el mismo reporte FO-CC-03
        recibos = ReciboMaterial.query.filter_by(reporte_focc03=reporte).order_by(ReciboMaterial.fecha_creacion).all()
        
        if not recibos:
            flash('No hay recibos para el reporte especificado', 'danger')
            return redirect(url_for('index'))
        
        # Crear libro de Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte FO-CC-03"
        
        # Estilos
        header_font = Font(color="FFFFFF", bold=True)
        header_fill = PatternFill(start_color="DC0000", end_color="DC0000", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título del reporte
        ws['D1'] = 'REPORTE DE INSPECCIÓN MATERIA PRIMA FO-CC-03'
        ws.merge_cells('D1:G1')
        ws['D1'].font = Font(bold=True, size=14)
        ws['D1'].alignment = Alignment(horizontal='center')
        
        # Información del reporte
        ws['A3'] = 'Reporte:'
        ws['B3'] = reporte
        ws['A4'] = 'Fecha:'
        ws['B4'] = datetime.datetime.now().strftime('%d/%m/%Y')
        
        # Encabezados
        headers = [
            'Item', 'Quantity', 'Material Description', 'Grade',
            'Plate/Coil No.', 'Heat/Batch', 'Certificate No.', 'ID',
            'Result', 'Remission'
        ]
        
        # Aplicar encabezados (fila 6)
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Llenar con datos
        for idx, recibo in enumerate(recibos, 1):
            row = idx + 6
            ws.cell(row=row, column=1, value=idx)  # Item
            ws.cell(row=row, column=2, value=recibo.cantidad)  # Quantity
            ws.cell(row=row, column=3, value=recibo.descripcion_material)  # Material Description
            ws.cell(row=row, column=4, value=recibo.grado_acero)  # Grade
            ws.cell(row=row, column=5, value=recibo.num_placa)  # Plate/Coil No.
            ws.cell(row=row, column=6, value=recibo.num_colada)  # Heat/Batch
            ws.cell(row=row, column=7, value=recibo.num_certificado)  # Certificate No.
            ws.cell(row=row, column=8, value=recibo.ot)  # ID
            ws.cell(row=row, column=9, value=recibo.estatus)  # Result
            ws.cell(row=row, column=10, value=recibo.num_remision)  # Remission
            
            # Aplicar bordes a todas las celdas
            for col in range(1, 11):
                ws.cell(row=row, column=col).border = thin_border
        
        # Firmas al final
        row = len(recibos) + 9
        ws.cell(row=row, column=2, value='Elaboró:')
        ws.cell(row=row, column=6, value='Revisó:')
        ws.cell(row=row, column=9, value='Autorizó:')
        
        row += 3
        ws.cell(row=row, column=2, value='_________________')
        ws.cell(row=row, column=6, value='_________________')
        ws.cell(row=row, column=9, value='_________________')
        
        row += 1
        ws.cell(row=row, column=2, value='Nombre y Firma')
        ws.cell(row=row, column=6, value='Nombre y Firma')
        ws.cell(row=row, column=9, value='Nombre y Firma')
        
        # Ajustar anchos de columna - Corregido para evitar error con celdas combinadas
        for column_cells in ws.columns:
            length = 0
            column = None
            for cell in column_cells:
                if hasattr(cell, 'column_letter'):  # Verificar si es una celda normal
                    column = cell.column_letter
                    if cell.value:
                        length = max(length, len(str(cell.value)))
            
            if length > 0 and column:  # Solo ajustar si tenemos una columna válida
                adjusted_width = (length + 2)
                ws.column_dimensions[column].width = adjusted_width
            
        # Crear buffer para el archivo
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Nombre del archivo
        filename = f"Reporte_FOCC03_{reporte}_{datetime.datetime.now().strftime('%Y-%m-%d')}.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'Error al exportar reporte: {str(e)}', 'danger')
        return redirect(url_for('index'))

@app.route('/importar_sqlserver', methods=['POST'])
def importar_sqlserver():
    response = {
        'status': 'error',
        'message': 'No se pudo procesar la solicitud',
        'logs': []
    }
    
    def add_log(message):
        response['logs'].append(message)
    
    try:
        # Verificar IDs de recibos
        ids = request.json.get('ids', [])
        if not ids:
            raise Exception('No se especificaron IDs para importar')
        
        # Validar IDs
        ids = [int(id) for id in ids if int(id) > 0]
        if not ids:
            raise Exception('No se especificaron IDs válidos')
        
        add_log("Iniciando proceso de importación")
        
        # Conexión a SQL Server de producción
        conn_prod = get_sqlserver_prod_conn()
        add_log("Conexión a SQL Server de producción establecida")
        
        # Establecer conexión con Firebird vía SSH
        firebird_result = get_firebird_conn()
        if not firebird_result['status']:
            raise Exception(f'Error al conectar con Firebird: {firebird_result["message"]}')
        
        fb_conn = firebird_result['conn']
        tunnel = firebird_result['tunnel']
        add_log("Conexión a Firebird establecida")
        
        # Procesar cada recibo
        recibos_procesados = 0
        errores = 0
        
        for id in ids:
            try:
                # Obtener datos del recibo
                recibo = ReciboMaterial.query.get(id)
                if not recibo:
                    add_log(f"Recibo ID {id} no encontrado")
                    errores += 1
                    continue
                
                add_log(f"Procesando recibo ID {id}: {recibo.descripcion_material}")
                
                # 1. Obtener DOCTO_CM_ID de Firebird
                if not recibo.orden_compra:
                    add_log(f"Recibo ID {id} no tiene orden de compra")
                    errores += 1
                    continue
                
                cursor_fb = fb_conn.cursor()
                cursor_fb.execute('SELECT "DOCTO_CM_ID" FROM "DOCTOS_CM" WHERE "FOLIO" = ? AND "TIPO_DOCTO" = ?', 
                                 (recibo.orden_compra, 'O'))
                docto_cm_id_row = cursor_fb.fetchone()
                
                if not docto_cm_id_row:
                    add_log(f"No se encontró DOCTO_CM_ID para la orden de compra {recibo.orden_compra}")
                    errores += 1
                    continue
                
                docto_cm_id = docto_cm_id_row[0]
                add_log(f"DOCTO_CM_ID encontrado: {docto_cm_id}")
                
                # 2. Insertar en tb_recibomtlcalidad
                cursor_prod = conn_prod.cursor()
                cursor_prod.execute("INSERT INTO tb_recibomtlcalidad (idOrdenCompra, lote) VALUES (?, ?)",
                                   (docto_cm_id, 1))
                conn_prod.commit()
                
                # Obtener el ID generado
                cursor_prod.execute("SELECT @@IDENTITY")
                idrecibo = cursor_prod.fetchval()
                
                if not idrecibo:
                    add_log("Error al obtener el ID del recibo insertado")
                    conn_prod.rollback()
                    errores += 1
                    continue
                
                add_log(f"Registro insertado en tb_recibomtlcalidad con ID: {idrecibo}")
                
                # 3. Obtener ARTICULO_ID y CLAVE_ARTICULO de Firebird
                if not recibo.descripcion_material:
                    add_log(f"Recibo ID {id} no tiene descripción de material")
                    conn_prod.rollback()
                    errores += 1
                    continue
                
                # a) Buscar ARTICULO_ID
                cursor_fb.execute('SELECT "ARTICULO_ID" FROM "ARTICULOS" WHERE "NOMBRE" = ?', 
                                 (recibo.descripcion_material,))
                articulo_id_row = cursor_fb.fetchone()
                
                if not articulo_id_row:
                    add_log(f"No se encontró ARTICULO_ID para la descripción: {recibo.descripcion_material}")
                    conn_prod.rollback()
                    errores += 1
                    continue
                
                articulo_id = articulo_id_row[0]
                add_log(f"ARTICULO_ID encontrado: {articulo_id}")
                
                # b) Buscar CLAVE_ARTICULO
                cursor_fb.execute('SELECT "CLAVE_ARTICULO" FROM "PRECIOS_COMPRA" WHERE "ARTICULO_ID" = ?', 
                                 (articulo_id,))
                clave_articulo_row = cursor_fb.fetchone()
                
                if not clave_articulo_row:
                    add_log(f"No se encontró CLAVE_ARTICULO para ARTICULO_ID: {articulo_id}")
                    conn_prod.rollback()
                    errores += 1
                    continue
                
                clave_articulo = clave_articulo_row[0]
                add_log(f"CLAVE_ARTICULO encontrado: {clave_articulo}")
                
                # 4. Insertar en tb_recibomtlcalidaddetalle
                cursor_prod.execute("""
                    INSERT INTO tb_recibomtlcalidaddetalle (
                        idrecibo, idProducto, descripcion, cantidad, idClave, 
                        usuarioalta, fechaalta, lote, comentarios, idestatus, 
                        iduom, idprocedencia
                    ) VALUES (?, ?, ?, ?, ?, ?, GETDATE(), ?, ?, ?, ?, ?)
                """, (
                    idrecibo,
                    clave_articulo,
                    recibo.descripcion_material,
                    recibo.cantidad,
                    recibo.orden_compra,
                    'jennifert',
                    recibo.idcode,
                    recibo.cliente,
                    9,
                    21,
                    recibo.procedencia
                ))
                
                conn_prod.commit()
                add_log("Registro insertado en tb_recibomtlcalidaddetalle")
                
                # 5. Actualizar el ID de orden de compra en la tabla local
                recibo.idordencompra = docto_cm_id
                db.session.commit()
                
                recibos_procesados += 1
                add_log(f"Recibo ID {id} procesado correctamente")
                
            except Exception as e:
                add_log(f"Error en recibo ID {id}: {str(e)}")
                if 'cursor_prod' in locals() and 'conn_prod' in locals():
                    conn_prod.rollback()
                errores += 1
                continue
        
        # Cerrar conexiones
        if 'conn_prod' in locals():
            conn_prod.close()
        
        if 'fb_conn' in locals():
            fb_conn.close()
        
        if 'tunnel' in locals() and tunnel.is_active:
            tunnel.close()
        
        # Respuesta exitosa
        if recibos_procesados > 0:
            response['status'] = 'success'
            response['message'] = f"Se importaron {recibos_procesados} recibos correctamente" + (
                f". Hubo {errores} errores." if errores > 0 else "."
            )
        else:
            response['message'] = f"No se pudo importar ningún recibo. Hubo {errores} errores."
        
    except Exception as e:
        response['message'] = f'Error general: {str(e)}'
        # Cerrar conexiones en caso de error
        if 'conn_prod' in locals():
            conn_prod.close()
        if 'fb_conn' in locals():
            fb_conn.close()
        if 'tunnel' in locals() and tunnel.is_active:
            tunnel.close()
    
    return jsonify(response)
@app.route('/buscar_articulos_por_oc/<orden_compra>', methods=['GET'])
def buscar_articulos_por_oc(orden_compra):
    try:
        # Establecer conexión con Firebird
        firebird_result = get_firebird_conn()
        if not firebird_result['status']:
            return jsonify({'status': 'error', 'message': f'Error al conectar con Firebird: {firebird_result["message"]}'})
        
        fb_conn = firebird_result['conn']
        tunnel = firebird_result['tunnel']
        
        try:
            cursor = fb_conn.cursor()
            
            # Primero obtener el DOCTO_CM_ID de la orden de compra
            cursor.execute('SELECT "DOCTO_CM_ID" FROM "DOCTOS_CM" WHERE "FOLIO" = ? AND "TIPO_DOCTO" = ?', 
                          (orden_compra, 'O'))
            docto_cm_id_row = cursor.fetchone()
            
            if not docto_cm_id_row:
                return jsonify({'status': 'error', 'message': f'No se encontró la orden de compra {orden_compra}'})
            
            docto_cm_id = docto_cm_id_row[0]
            
            # Obtener los detalles de la orden de compra
            cursor.execute("""
                SELECT DET."CLAVE_ARTICULO", DET."ARTICULO_ID", ART."NOMBRE" as ARTICULO,
                       DET."UNIDADES", DET."PRECIO_UNITARIO", PROV."NOMBRE" as PROVEEDOR
                FROM "DOCTOS_CM_DET" DET
                INNER JOIN "ARTICULOS" ART ON ART."ARTICULO_ID" = DET."ARTICULO_ID"
                INNER JOIN "DOCTOS_CM" OC ON OC."DOCTO_CM_ID" = DET."DOCTO_CM_ID"
                INNER JOIN "PROVEEDORES" PROV ON PROV."PROVEEDOR_ID" = OC."PROVEEDOR_ID"
                WHERE DET."DOCTO_CM_ID" = ?
            """, (docto_cm_id,))
            
            articulos = []
            for row in cursor.fetchall():
                articulos.append({
                    'clave_articulo': row[0],
                    'articulo_id': row[1],
                    'descripcion': row[2],
                    'unidades': row[3],
                    'precio_unitario': row[4],
                    'proveedor': row[5]
                })
            
            return jsonify({
                'status': 'success',
                'articulos': articulos,
                'docto_cm_id': docto_cm_id
            })
            
        finally:
            fb_conn.close()
            if tunnel.is_active:
                tunnel.close()
    
    except Exception as e:
        if 'tunnel' in locals() and tunnel.is_active:
            tunnel.close()
        return jsonify({'status': 'error', 'message': str(e)})
if __name__ == '__main__':
    # Crear tablas si no existen
    with app.app_context():
        db.create_all()
    app.run(debug=True, host='0.0.0.0', port=5000)